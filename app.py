import os
import json
# sys.path.insert(0, '/path/to/website/app')
from time import gmtime, strftime, sleep
from datetime import datetime
from random import randint
import xml.etree.ElementTree as ET
from soapproxy import soapproxy
import flask_mysqldb
mysql = flask_mysqldb.MySQL()
import flask 
from flask import Flask, render_template, redirect, url_for, request,session,jsonify, abort
from openpyxl import load_workbook
from werkzeug import secure_filename

from clustering import *
from updateDB import *
import numpy as np 
import requests
from scipy.optimize import linear_sum_assignment
from flask_restful import reqparse, abort, Api, Resource
from requests import put, get

app = Flask(__name__)
api = Api(app)

app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = 'tAxzg(bMEWK'
app.config['MYSQL_DB'] = 'dispatch_db'
app.config['MYSQL_HOST'] = 'localhost'
app.config['UPLOAD_FOLDER']='./upload'
ALLOWED_EXTENSIONS = set(['xlsx','xls','txt', 'pdf', 'jpg', 'jpeg', 'gif'])
mysql.init_app(app)
app.secret_key = os.urandom(24)

def allowed_filename(filename):
    return '.' in filename and filename.rsplit('.',1)[1] in ALLOWED_EXTENSIONS

@app.route("/", methods=['GET', 'POST'])
def main():
    error = None
    if request.method == 'POST':
        username=request.form['uname']

        cursor = mysql.connection.cursor()
        cursor.execute("SELECT Password,appid, Id from dispatch_db.company where Username='" + username + "'")
        data=cursor.fetchone()
        
        session['username'] = username
        session['pswd'] = data[0]
        session['appid'] = data[1]
        session['Id'] = data[2]

        mysoapproxy=soapproxy(username,data[0],data[1])
        if request.form['uname'] != username or request.form['pwd'] != data[0]:
            error = 'Invalid Credentials. Please try again.'
        else:
            session['logged_in'] = True
            return redirect(url_for('home'))
    return render_template('index.html')

@app.route("/home")
def home():
    username=session['username']
    appid=session['appid']
    SITE_ROOT = os.path.realpath(os.path.dirname(__file__))
    logfile = os.path.join(SITE_ROOT, 'static', 'chatfile/log232.html')
    f=open(logfile,"r")
    content=f.read()
    print(content)
    f.close()
    print(session['logged_in'])
    return render_template('home.html',content=content)

@app.route("/markcar",methods=["GET", "POST"])
def markcar():
    if request.method == "GET":
        cpnId = str(session['Id'])
        mysoapproxy=soapproxy('TianWeiAdmin','TianWei123','232')
        device = mysoapproxy.getUsersInUserTag('All Vehicles')
       # Should comment out this line later
        db = mysql.connection
        cursor = db.cursor()
        for vehicle in device:
            # Check if this vehicle exist in the db yet
            cursor.execute("SELECT * FROM vehiclelist WHERE cpnId='" + cpnId + "'")
            data = cursor.fetchall()
            # Not exist
            if len( data ) == 0:
                sql = "INSERT INTO vehiclelist (cpnId, plate, status, longitude, latitude) VALUES (%s, %s, %s, %s, %s)"
                data = cursor.execute(sql, (session['Id'], vehicle[0], 'idle', vehicle[1], vehicle[2]))
                db.commit()
    session['numdev'] = len(device)
    return jsonify(device)


@app.route("/uploadsinglejob",methods=["GET", "POST"])
def uploadsinglejob(): 
    from datetime import datetime
    cpnId = str(session['Id'])
    db = mysql.connection
    cursor = db.cursor()
    if request.method == "POST":
        jobdate = datetime.now().strftime( "%Y-%m-%d %I:%M:%S" ) 
        startDt, endDt = jobdate[:-8] + "00:00:00", jobdate[:-8] + "11:59:00"

        if request.form['jobaddress'] != None:
            address = request.form['jobaddress']    
            description = request.form['jobdesc']
            insertJobToDB(db, address, jobdate, cpnId, description)
            allJobs = 'success'
    return jsonify(allJobs)

@app.route("/uploadfile",methods=["GET", "POST"])
def uploadfile():
    from datetime import datetime
    cpnId = str(session['Id'])
    db = mysql.connection
    cursor = db.cursor()
    if request.method == "POST":
        jobdate = datetime.now().strftime( "%Y-%m-%d %I:%M:%S" ) 
        startDt, endDt = jobdate[:-8] + "00:00:00", jobdate[:-8] + "11:59:00"
        output, addressList, descriptionList = [], [], []

        if request.files['file'] != None:
            file = request.files['file']
            if file and allowed_filename(file.filename):
                filename = secure_filename(file.filename)
                file.save(os.path.join('C:/Apache24/htdocs/app/upload',filename))
                SITE_ROOT = os.path.realpath(os.path.dirname(__file__))
                uploadfile = os.path.join('C:/Apache24/htdocs/app/upload',filename)
                filepath = 'C:/Apache24/htdocs/app/upload/' + filename
                # Check if file already exist in the system
                if os.path.isfile(uploadfile):
                    wb=load_workbook(uploadfile)
                    job_num=wb['Sheet1'].max_row
                    for i in range(1, job_num+1):
                        addressList.append( wb['Sheet1'].cell(row=i,column=1).value )
                        descriptionList.append( wb['Sheet1'].cell(row=i,column=2).value )

                    insertFileToDB(db, addressList, jobdate, cpnId, descriptionList) 
                    allJobs = 'success'
    return jsonify(allJobs)

@app.route("/assignJob", methods = ["GET", "POST"])
def assignJob(): 
    db = mysql.connection
    cpnId = str(session['Id'])
    jobdate = datetime.now().strftime( "%Y-%m-%d %I:%M:%S" ) 
    startDt, endDt = jobdate[:-8] + "00:00:00", jobdate[:-8] + "11:59:00"
    #Query all "Unassigned" jobs
    unassigned_jobs = getTodayUnassignedJobs(db, cpnId, startDt, endDt)  

    numdev = session['numdev']
    cursor = db.cursor()
    job_list = func(unassigned_jobs, numdev) # job_list = [ [address, lat, long, routeInd, description], [route2], [route3], ...]   
    result = ['Process jobs successfully!!!', job_list]

  # get all avaialable vehicles to assign jobs
    cursor.execute("SELECT plate, longitude, latitude FROM dispatch_db.vehiclelist WHERE cpnId='" + cpnId + "'")
    vehiclelist = cursor.fetchall()
    routeAssignment = {}
    distBwVehJob = [ [None]*len(vehiclelist) ] * len(job_list)

    #assign jobs to the closet vehicle
    for routeInd, route in enumerate(job_list):   
        for vehInd, vehicle in enumerate(vehiclelist):
            distBwVehJob[routeInd][vehInd] = distance(vehicle[2], vehicle[1], route[0][2], route[0][1])
    row_ind, chosenVehInd = linear_sum_assignment(distBwVehJob)
    chosenVeh = [ vehiclelist[i][0] for i in chosenVehInd]

    for routeInd, route in enumerate(distBwVehJob):
        routeAssignment[ chosenVeh[routeInd] ] = job_list[routeInd]  # get the plate number of chosen vehicle
        for i in range(len(job_list[routeInd])):
            sql = "UPDATE dispatch_db.jobs SET plate=%s, status = %s, routeInd=%s WHERE jobId = %s"
            cursor.execute(sql, (chosenVeh[routeInd] , 'Assigned', i, job_list[routeInd][i][3])) 
            db.commit()

    return jsonify(result)

@app.route('/getAvailableJobs', methods=['GET', 'POST'])
def getAvailableJobs():
    cpnId = str(session['Id'])
    jobdate = datetime.now().strftime( "%Y-%m-%d %I:%M:%S" ) 
    startDt, endDt = jobdate[:-8] + "00:00:00", jobdate[:-8] + "11:59:00"
    db = mysql.connection
    cursor = db.cursor()
    cursor.execute("SELECT address, latitude, longitude, routeInd, description FROM dispatch_db.jobs WHERE plate = 'Unassigned' AND  status= 'Unassigned' AND jobdate BETWEEN '"+startDt+"' AND '"+endDt+"'")
    unassigned_jobs = cursor.fetchall()
    print('unassigned' , unassigned_jobs)
    job_list = [ unassigned_jobs ]
    cursor.execute("SELECT DISTINCT plate from dispatch_db.jobs WHERE cpnId='" + cpnId + "' AND status = 'Assigned' ")
    vehiclelist = cursor.fetchall()
    for p in vehiclelist:
        print('p',p)
        cursor.execute("SELECT address, latitude, longitude, routeInd, description FROM dispatch_db.jobs WHERE plate='"+p[0]+"' AND status!= 'Finished' AND jobdate BETWEEN '"+startDt+"' AND '"+endDt+"'")    #result = cursor.fetchall()
        job = cursor.fetchall()
        print('job', job)
        job_list.append(list(job))
    result = ['Process jobs successfully!!!', job_list]
    print('getAvailableJobs', result)
    return jsonify(result)


@app.route('/finished', methods=['GET'])
def finished():
    jobId = request.form['jobId']
    sql = "UPDATE dispatch_db.jobs SET status = %s WHERE ID = %s"
    cursor.execute(sql, ('Finished', jobId)) 
    db.commit()

@app.route("/robot",methods=["GET", "POST"])
def robot():
    if request.method == "POST":
      msg=request.form['text']
      SITE_ROOT = os.path.realpath(os.path.dirname(__file__))
      logfile = os.path.join(SITE_ROOT, 'static', 'chatfile/log232.html')
      f=open(logfile,"a")
      f.write(msg+"<br/>") 
      f.close()
      print(msg)
      remsg=json.dumps({'user':msg, 'robot':'ok'})
      return  jsonify(remsg)

@app.route('/login', methods=['GET'])
def get_tasks():
    return jsonify({'logins': logins})


@app.route('/login/<string:username>', methods=['GET'])
def get_task(task_id):
    login = [login for login in logins if login['username'] == username]
    if len(login) == 0:
        abort(404)
    return jsonify({'login': login[0]})

from flask import make_response
@app.errorhandler(404)
def not_found(error):
    return make_response(jsonify({'error': 'Not found'}), 404)

@app.route('/login', methods=['POST'])
def create_task():
    if not request.form:
        abort(400)
    username =  request.form['username']
    password = request.form['password']

    db = mysql.connection
    cursor = db.cursor()
    cursor.execute("SELECT plate from dispatch_db.company where Username='" + username + "' AND Password='"+password+"'")

    plate = cursor.fetchone()
    if len(plate) == 0:
        login = {"response": "fail", 
                "plate": ""}
    else:
        login = {"response": "success", 
                "plate": plate[0]}

    return jsonify(login), 201

@app.route('/driver', methods=['POST'])
def get_jobs():
    if request.method == "POST":
        plate=request.form['plate']

        db = mysql.connection
        cursor = db.cursor()
        cursor.execute("SELECT * from dispatch_db.jobs where plate = '" + plate + "' order by routeInd")
        data = cursor.fetchall()
        if len(data) == 0:
            login = {"response": "fail", 
                    "plate": ""}
        else:
            login = {"response": "success", 
                    "plate": data}
        return jsonify(login), 201
    

if __name__ == "__main__":
    
    app.run()



